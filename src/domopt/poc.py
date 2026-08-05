"""Strict adapter for the readable Nestle proof-of-concept challenge bundle.

The adapter never writes raw operational data. It converts the supplied files
in memory to the canonical DOM contract and exposes aggregate audits suitable
for reports. Recommendation outputs are benchmark evidence, not optimizer labels.
"""

from __future__ import annotations

import hashlib
import math
import zipfile
from collections.abc import Iterable
from dataclasses import dataclass, replace
from functools import cache
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from pypdf import PdfReader

from .data import DataValidationError, normalize_problem_data
from .penalties import order_penalty
from .schemas import ASSUMPTION_VERSION, SCHEMA_VERSION, ProblemData

POC_FILENAMES = {
    "equations": "DOM Equations.docx",
    "example": "Example.xlsx",
    "order_output": "Output_order_level_data(3).csv",
    "sku_output": "output_order_sku_level_data(3).csv",
    "inventory": "input_capacity_planning.csv",
    "orders": "input_order data.csv",
    "shipping": "input_shipping_cost_data.csv",
    "dock": "input_dock_capacity.csv",
    "throughput": "input_throughput_capacity.csv",
    "challenge": "Nestle - WISER Quantum Challenge [SHARED](2).pdf",
}


class PocDataError(DataValidationError):
    """Raised when a supplied POC artifact is missing, unreadable, or inconsistent."""


@dataclass(frozen=True)
class PocConfig:
    """Business assumptions that are explicit, tested, and notebook-visible."""

    focus_only: bool = True
    protection_days: int = 5
    miles_per_lead_day: float = 500.0
    min_divert_improvement_fraction: float = 0.05
    min_divert_improvement_cases: int = 100
    enforce_assignment_group: bool = True
    throughput_headroom_fraction: float | None = None
    holidays: tuple[str, ...] = ()
    pareto_prune: bool = True

    def validate(self) -> None:
        if self.protection_days < 0:
            raise ValueError("protection_days must be nonnegative")
        if self.miles_per_lead_day <= 0:
            raise ValueError("miles_per_lead_day must be positive")
        if not 0 <= self.min_divert_improvement_fraction <= 1:
            raise ValueError("min_divert_improvement_fraction must be in [0, 1]")
        if self.min_divert_improvement_cases < 0:
            raise ValueError("min_divert_improvement_cases must be nonnegative")
        if (
            self.throughput_headroom_fraction is not None
            and self.throughput_headroom_fraction < 0
        ):
            raise ValueError("throughput_headroom_fraction must be nonnegative")


def _paths(bundle_dir: str | Path) -> dict[str, Path]:
    root = Path(bundle_dir)
    return {key: root / name for key, name in POC_FILENAMES.items()}


def _read_csv(path: Path, *, dtype: dict[str, str] | None = None) -> pd.DataFrame:
    try:
        frame = pd.read_csv(path, dtype=dtype, low_memory=False)
    except Exception as error:  # pragma: no cover - parser messages vary
        raise PocDataError(f"Unreadable CSV {path.name}: {error}") from error
    if frame.empty:
        raise PocDataError(f"CSV contains no data rows: {path.name}")
    return frame


def audit_poc_bundle(bundle_dir: str | Path) -> pd.DataFrame:
    """Read every supplied artifact and fail immediately on the first bad file."""

    paths = _paths(bundle_dir)
    rows: list[dict[str, object]] = []
    for key, path in paths.items():
        if not path.is_file() or path.stat().st_size == 0:
            raise PocDataError(f"Required challenge file is missing or empty: {path}")
        suffix = path.suffix.lower()
        record: dict[str, object] = {
            "role": key,
            "filename": path.name,
            "bytes": path.stat().st_size,
            "readable": True,
        }
        try:
            if suffix == ".csv":
                frame = pd.read_csv(path, low_memory=False)
                if frame.empty:
                    raise ValueError("no data rows")
                record.update(rows=len(frame), columns=len(frame.columns))
            elif suffix == ".xlsx":
                workbook = load_workbook(path, read_only=True, data_only=False)
                if not workbook.sheetnames:
                    raise ValueError("no worksheets")
                record.update(
                    rows=max(workbook[name].max_row for name in workbook.sheetnames),
                    columns=max(
                        workbook[name].max_column for name in workbook.sheetnames
                    ),
                )
                workbook.close()
            elif suffix == ".docx":
                with zipfile.ZipFile(path) as archive:
                    bad = archive.testzip()
                    if bad is not None or "word/document.xml" not in archive.namelist():
                        raise ValueError(f"invalid OOXML member: {bad}")
                    xml = archive.read("word/document.xml")
                record.update(rows=xml.count(b"<w:p"), columns=np.nan)
            elif suffix == ".pdf":
                reader = PdfReader(path)
                if len(reader.pages) == 0:
                    raise ValueError("no pages")
                for page in reader.pages:
                    page.extract_text()
                record.update(rows=len(reader.pages), columns=np.nan)
            else:
                raise ValueError(f"unsupported file type {suffix}")
        except Exception as error:
            raise PocDataError(f"Unreadable challenge file {path.name}: {error}") from error
        rows.append(record)
    return pd.DataFrame(rows)


def _id(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    return text[:-2] if text.endswith(".0") and text[:-2].isdigit() else text


def _numeric(series: pd.Series, *, fill: float = 0.0) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").fillna(fill).astype(float)


def _case_demand(source: pd.DataFrame) -> pd.Series:
    planning = _numeric(source["OrderedQty_converted"])
    per_case = pd.to_numeric(source["ProductPlanningUnitsPerCase"], errors="coerce")
    cases_per_pallet = _numeric(source["ProductCasesPerPallet"])
    cases = np.where(
        per_case.notna() & (per_case > 0),
        planning / per_case,
        planning * cases_per_pallet,
    )
    rounded = np.rint(cases)
    if not np.allclose(cases, rounded, atol=5e-4):
        bad = source.loc[np.abs(cases - rounded) > 5e-4].head()
        raise PocDataError(
            "Planning-unit conversion does not produce integer cases for rows "
            f"{bad.index.tolist()}"
        )
    return pd.Series(rounded.astype(np.int64), index=source.index)


def _latest_working_pgi(
    requested_delivery: pd.Timestamp,
    lead_days: int,
    holidays: set[pd.Timestamp],
) -> pd.Timestamp:
    date = pd.Timestamp(requested_delivery).normalize() - pd.Timedelta(days=lead_days)
    while date.weekday() >= 5 or date in holidays:
        date -= pd.Timedelta(days=1)
    return date


def _bundle_hash(paths: Iterable[Path]) -> str:
    digest = hashlib.sha256()
    for path in sorted(paths, key=lambda item: item.name):
        digest.update(path.name.encode())
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
    return digest.hexdigest()


def _build_source_tables(bundle_dir: str | Path) -> dict[str, pd.DataFrame]:
    paths = _paths(bundle_dir)
    ids = {
        "Group_Flag": str,
        "Plant": str,
        "MaterialNumber": str,
        "ZipCode": str,
        "LoadNumber": str,
    }
    return {
        "orders": _read_csv(paths["orders"], dtype=ids),
        "inventory": _read_csv(
            paths["inventory"], dtype={"LocationID": str, "MaterialID": str}
        ),
        "shipping": _read_csv(
            paths["shipping"], dtype={"Plant": str, "TargetZip": str}
        ),
        "dock": _read_csv(paths["dock"], dtype={"Plant": str}),
        "throughput": _read_csv(paths["throughput"], dtype={"Plant": str}),
        "order_output": _read_csv(
            paths["order_output"],
            dtype={"SalesDocument/GroupingIndicator": str},
        ),
        "sku_output": _read_csv(
            paths["sku_output"],
            dtype={
                "SalesDocument/GroupingIndicator": str,
                "MaterialNumber": str,
            },
        ),
    }


def _canonical_lines(source: pd.DataFrame) -> pd.DataFrame:
    frame = source.copy()
    frame["order_id"] = frame["Group_Flag"].map(_id)
    frame["sku_id"] = frame["MaterialNumber"].map(_id)
    frame["demand_cases"] = _case_demand(frame)
    if (frame["demand_cases"] <= 0).any():
        raise PocDataError("Every challenge order line must have positive case demand")
    revenue = _numeric(frame["Order_SKU_Revenue"])
    frame["unit_value"] = revenue / frame["demand_cases"]
    penalty_fraction = _numeric(frame["Penaltyforpotentialcuts"])
    frame["penalty_per_unfilled_case"] = frame["unit_value"] * penalty_fraction
    frame["cases_per_pallet"] = np.rint(
        _numeric(frame["ProductCasesPerPallet"])
    ).astype(int)
    frame["unit_weight"] = _numeric(frame["OrderedWeight"]) / frame["demand_cases"]
    frame["unit_volume"] = _numeric(frame["OrderedVolume"]) / frame["demand_cases"]
    frame["forecast_required"] = True
    columns = [
        "order_id",
        "sku_id",
        "demand_cases",
        "unit_value",
        "penalty_per_unfilled_case",
        "cases_per_pallet",
        "unit_weight",
        "unit_volume",
        "forecast_required",
    ]
    lines = frame[columns].copy()
    if lines.duplicated(["order_id", "sku_id"]).any():
        raise PocDataError("Input order data duplicate an order-SKU key")
    return lines


def _first_value(group: pd.DataFrame, column: str) -> object:
    values = group[column].dropna().unique()
    if len(values) > 1:
        raise PocDataError(f"Order {group.name!r} has multiple values for {column}")
    return values[0] if len(values) else np.nan


def _canonical_orders(source: pd.DataFrame) -> pd.DataFrame:
    frame = source.copy()
    frame["order_id"] = frame["Group_Flag"].map(_id)
    frame["default_dc"] = frame["Plant"].map(_id)
    frame["default_pgi_date"] = pd.to_datetime(
        frame["transportationplanningdate"], errors="coerce", format="mixed"
    )
    frame["requested_delivery_date"] = pd.to_datetime(
        frame["RequestedDeliveryDate"], errors="coerce", format="mixed"
    )
    if frame[["default_pgi_date", "requested_delivery_date"]].isna().any().any():
        raise PocDataError("Order input contains invalid PGI or requested-delivery dates")
    frame["assignment_group"] = frame["LoadNumber"].map(_id)
    missing_load = frame["assignment_group"].eq("")
    frame.loc[missing_load, "assignment_group"] = (
        "ORDER::" + frame.loc[missing_load, "order_id"]
    )

    records: list[dict[str, object]] = []
    invariant = [
        "default_dc",
        "default_pgi_date",
        "requested_delivery_date",
        "assignment_group",
        "DeliveryPriority",
        "IsTopCust",
        "FillRateThreshold",
        "FixedPenalty",
        "FixedPenaltyPerSKU",
        "MinimumPenalty",
        "MaximumPenalty",
        "ZipCode",
    ]
    for order_id, group in frame.groupby("order_id", sort=False):
        group.name = order_id
        values = {column: _first_value(group, column) for column in invariant}
        records.append(
            {
                "order_id": str(order_id),
                "default_dc": _id(values["default_dc"]),
                "default_pgi_date": pd.Timestamp(values["default_pgi_date"]),
                "requested_delivery_date": pd.Timestamp(
                    values["requested_delivery_date"]
                ),
                "assignment_group": _id(values["assignment_group"]),
                "priority": int(float(values["DeliveryPriority"])),
                "is_top_customer": str(values["IsTopCust"]).upper() == "Y",
                "penalty_threshold_fraction": 0.0
                if pd.isna(values["FillRateThreshold"])
                else float(values["FillRateThreshold"]),
                "penalty_fixed": 0.0
                if pd.isna(values["FixedPenalty"])
                else float(values["FixedPenalty"]),
                "penalty_per_cut_sku": 0.0
                if pd.isna(values["FixedPenaltyPerSKU"])
                else float(values["FixedPenaltyPerSKU"]),
                "penalty_minimum": 0.0
                if pd.isna(values["MinimumPenalty"])
                else float(values["MinimumPenalty"]),
                "penalty_maximum": 0.0
                if pd.isna(values["MaximumPenalty"])
                else float(values["MaximumPenalty"]),
                "destination_zip": _id(values["ZipCode"]),
                "default_fillable_cases": 0,
                "min_divert_improvement_fraction": 0.05,
            }
        )
    return pd.DataFrame(records)


def _filter_focus(
    source: pd.DataFrame,
    orders: pd.DataFrame,
    lines: pd.DataFrame,
    *,
    focus_only: bool,
    enforce_assignment_group: bool,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    if not focus_only:
        return orders, lines
    source_ids = source["Group_Flag"].map(_id)
    focus = set(source_ids.loc[source["IsInvAvail"].astype(str).str.upper() == "N"])
    if enforce_assignment_group:
        groups = set(
            orders.loc[orders["order_id"].isin(focus), "assignment_group"].astype(str)
        )
        focus = set(
            orders.loc[orders["assignment_group"].astype(str).isin(groups), "order_id"]
        )
    return (
        orders.loc[orders["order_id"].isin(focus)].reset_index(drop=True),
        lines.loc[lines["order_id"].isin(focus)].reset_index(drop=True),
    )


def _prepare_inventory(source: pd.DataFrame) -> pd.DataFrame:
    frame = source.copy()
    frame["dc_id"] = frame["LocationID"].map(_id)
    frame["sku_id"] = frame["MaterialID"].map(_id)
    frame["date"] = pd.to_datetime(frame["DATE"], errors="coerce")
    frame["available"] = _numeric(frame["Available_inventory"]).clip(lower=0)
    if frame["date"].isna().any():
        raise PocDataError("Inventory input contains invalid dates")
    if frame.duplicated(["dc_id", "sku_id", "date"]).any():
        raise PocDataError("Inventory input duplicates a DC-SKU-date key")
    return frame


def _candidate_and_resource_tables(
    orders: pd.DataFrame,
    lines: pd.DataFrame,
    inventory_source: pd.DataFrame,
    shipping_source: pd.DataFrame,
    dock_source: pd.DataFrame,
    throughput_source: pd.DataFrame,
    config: PocConfig,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    shipping = shipping_source.copy()
    shipping["dc_id"] = shipping["Plant"].map(_id)
    shipping["destination_zip"] = shipping["TargetZip"].map(_id)
    shipping["distance"] = _numeric(shipping["Distance"])
    shipping["shipping_cost"] = _numeric(shipping["Shipping_Cost"])
    if shipping.duplicated(["dc_id", "destination_zip"]).any():
        raise PocDataError("Shipping input duplicates a plant-target ZIP lane")
    lane_lookup = shipping.set_index(["dc_id", "destination_zip"])

    dock = dock_source.copy()
    dock["dc_id"] = dock["Plant"].map(_id)
    dock["date"] = pd.to_datetime(dock["Date"], errors="coerce", format="mixed")
    dock["remaining"] = _numeric(dock["Dock_Remaining"]).clip(lower=0)
    dock_lookup = dock.set_index(["dc_id", "date"])["remaining"].to_dict()

    in_scope_dcs = sorted(set(orders["default_dc"].astype(str)))
    inventory_skus = {
        dc_id: set(group["sku_id"].astype(str))
        for dc_id, group in inventory_source.groupby("dc_id", sort=False)
    }
    inventory_profiles = {
        (str(dc_id), str(sku_id)): (
            group.sort_values("date")["date"].to_numpy(dtype="datetime64[ns]"),
            group.sort_values("date")["available"].to_numpy(dtype=float),
        )
        for (dc_id, sku_id), group in inventory_source.groupby(
            ["dc_id", "sku_id"], sort=False
        )
    }

    @cache
    def available_atp(dc_id: str, sku_id: str, date: pd.Timestamp) -> int | None:
        profile = inventory_profiles.get((dc_id, sku_id))
        if profile is None:
            return None
        dates, amounts = profile
        start = np.datetime64(pd.Timestamp(date).to_datetime64())
        end = np.datetime64(
            (pd.Timestamp(date) + pd.Timedelta(days=config.protection_days)).to_datetime64()
        )
        covered = amounts[(dates >= start) & (dates <= end)]
        if len(covered) == 0:
            return None
        return max(0, math.floor(float(covered.min()) + 1e-9))
    lines_by_order = {
        str(order_id): group for order_id, group in lines.groupby("order_id", sort=False)
    }
    holiday_set = {pd.Timestamp(value).normalize() for value in config.holidays}
    records: list[dict[str, object]] = []

    for group_id, group_orders in orders.groupby("assignment_group", sort=False):
        group_orders = group_orders.sort_values("order_id", kind="mergesort")
        leader = str(group_orders.iloc[0]["order_id"])
        members = group_orders["order_id"].astype(str).tolist()
        destination_values = set(group_orders["destination_zip"].astype(str))
        rdd_values = set(group_orders["requested_delivery_date"])
        default_dc_values = set(group_orders["default_dc"].astype(str))
        default_pgi_values = set(group_orders["default_pgi_date"])
        if config.enforce_assignment_group and any(
            len(values) != 1
            for values in [
                destination_values,
                rdd_values,
                default_dc_values,
                default_pgi_values,
            ]
        ):
            raise PocDataError(
                f"Assignment group {group_id!r} has incompatible routing fields"
            )
        destination = min(destination_values)
        rdd = pd.Timestamp(min(rdd_values))
        default_dc = min(default_dc_values)
        default_pgi = pd.Timestamp(min(default_pgi_values))

        options: list[tuple[str, pd.Timestamp, int, float, float, bool]] = []
        default_key = (default_dc, destination)
        if default_key not in lane_lookup.index:
            raise PocDataError(
                f"Missing default shipping lane for assignment group {group_id!r}"
            )
        default_lane = lane_lookup.loc[default_key]
        options.append(
            (
                default_dc,
                default_pgi,
                math.ceil(float(default_lane["distance"]) / config.miles_per_lead_day),
                float(default_lane["distance"]),
                float(default_lane["shipping_cost"]),
                True,
            )
        )

        group_skus = set(
            lines.loc[lines["order_id"].isin(members), "sku_id"].astype(str)
        )
        for dc_id in in_scope_dcs:
            if dc_id == default_dc or (dc_id, destination) not in lane_lookup.index:
                continue
            if not group_skus <= inventory_skus.get(dc_id, set()):
                continue
            lane = lane_lookup.loc[(dc_id, destination)]
            distance = float(lane["distance"])
            lead_days = math.ceil(distance / config.miles_per_lead_day)
            pgi = _latest_working_pgi(rdd, lead_days, holiday_set)
            if dock_lookup.get((dc_id, pgi), 0.0) < 1.0:
                continue
            options.append(
                (
                    dc_id,
                    pgi,
                    lead_days,
                    distance,
                    float(lane["shipping_cost"]),
                    False,
                )
            )

        for dc_id, pgi, lead_days, distance, shipping_cost, is_default in options:
            option_id = f"{dc_id}::{pgi.date().isoformat()}"
            for order in group_orders.itertuples(index=False):
                order_id = str(order.order_id)
                estimated_fill = 0
                estimated_value = 0.0
                for line in lines_by_order[order_id].itertuples(index=False):
                    available = available_atp(dc_id, str(line.sku_id), pgi)
                    if available is None:
                        if is_default:
                            available = 0
                        else:
                            estimated_fill = -1
                            break
                    quantity = min(int(line.demand_cases), available)
                    estimated_fill += quantity
                    estimated_value += quantity * float(line.unit_value)
                if estimated_fill < 0:
                    break
                records.append(
                    {
                        "candidate_id": f"{order_id}__{option_id}",
                        "order_id": order_id,
                        "dc_id": dc_id,
                        "pgi_date": pgi,
                        "arrival_date": pgi + pd.Timedelta(days=lead_days),
                        "shipping_cost": shipping_cost if order_id == leader else 0.0,
                        "distance": distance,
                        "lead_time_days": lead_days,
                        "dock_units": (0.0 if is_default else 1.0)
                        if order_id == leader
                        else 0.0,
                        "is_default": is_default,
                        "eligible": True,
                        "group_option_id": option_id,
                        "estimated_fill_cases": estimated_fill,
                        "estimated_fulfilled_value": estimated_value,
                    }
                )

    candidates = pd.DataFrame(records)
    if candidates.empty:
        raise PocDataError("Candidate generation produced no eligible rows")

    if config.enforce_assignment_group:
        order_group = orders.set_index("order_id")["assignment_group"].to_dict()
        candidates["assignment_group"] = candidates["order_id"].map(order_group)
        complete: set[tuple[str, str]] = set()
        for group_id, group in candidates.groupby("assignment_group", sort=False):
            member_count = int(
                orders.loc[orders["assignment_group"] == group_id, "order_id"].nunique()
            )
            counts = group.groupby("group_option_id")["order_id"].nunique()
            complete.update(
                (str(group_id), str(option_id))
                for option_id in counts.index[counts == member_count]
            )
        candidates = candidates.loc[
            candidates.apply(
                lambda row: (str(row["assignment_group"]), str(row["group_option_id"]))
                in complete,
                axis=1,
            )
        ].drop(columns=["assignment_group"])

    candidate_dates = candidates[["dc_id", "pgi_date"]].drop_duplicates()
    relevant_skus = set(lines["sku_id"].astype(str))
    inventory_rows: list[dict[str, object]] = []
    for row in candidate_dates.itertuples(index=False):
        for sku_id in sorted(relevant_skus & inventory_skus.get(str(row.dc_id), set())):
            available = available_atp(
                str(row.dc_id), sku_id, pd.Timestamp(row.pgi_date)
            )
            if available is not None:
                inventory_rows.append(
                    {
                        "dc_id": str(row.dc_id),
                        "sku_id": sku_id,
                        "date": pd.Timestamp(row.pgi_date),
                        "cumulative_available_cases": available,
                    }
                )
    inventory = pd.DataFrame(inventory_rows).drop_duplicates(
        ["dc_id", "sku_id", "date"]
    )

    capacity_rows: list[dict[str, object]] = []
    for row in candidate_dates.itertuples(index=False):
        key = (str(row.dc_id), pd.Timestamp(row.pgi_date))
        if key in dock_lookup:
            capacity_rows.append(
                {
                    "dc_id": key[0],
                    "date": key[1],
                    "resource": "dock",
                    "capacity": max(0.0, float(dock_lookup[key])),
                    "unit": "incremental_loads",
                }
            )

    if config.throughput_headroom_fraction is not None:
        throughput = throughput_source.copy()
        throughput["dc_id"] = throughput["Plant"].map(_id)
        throughput["date"] = pd.to_datetime(
            throughput["transportationplanningdate"], errors="coerce"
        )
        for row in throughput.itertuples(index=False):
            if (str(row.dc_id), pd.Timestamp(row.date)) not in {
                (str(item.dc_id), pd.Timestamp(item.pgi_date))
                for item in candidate_dates.itertuples(index=False)
            }:
                continue
            for resource, source_column, unit in [
                ("case_pick", "util_case_picks", "loose_cases"),
                ("pallet_pick", "util_pallets", "full_pallets"),
            ]:
                observed = float(getattr(row, source_column))
                capacity_rows.append(
                    {
                        "dc_id": str(row.dc_id),
                        "date": pd.Timestamp(row.date),
                        "resource": resource,
                        "capacity": max(
                            0.0, observed * config.throughput_headroom_fraction
                        ),
                        "unit": unit,
                    }
                )
    capacities = pd.DataFrame(
        capacity_rows,
        columns=["dc_id", "date", "resource", "capacity", "unit"],
    ).drop_duplicates(["dc_id", "date", "resource"], keep="last")
    calendar = candidate_dates.rename(columns={"pgi_date": "date"}).copy()
    calendar["is_open"] = True
    return candidates.reset_index(drop=True), inventory, capacities, calendar


def _with_default_reference(problem: ProblemData) -> ProblemData:
    default_candidates = problem.candidates.loc[
        problem.candidates["is_default"].astype(bool)
    ]
    fill = default_candidates.set_index("order_id")["estimated_fill_cases"].to_dict()
    orders = problem.orders.copy()
    orders["default_fillable_cases"] = (
        orders["order_id"].map(fill).fillna(0).astype(int)
    )
    return replace(
        problem,
        orders=orders,
        metadata={
            **problem.metadata,
            "enforce_min_divert_improvement": True,
            "default_fill_reference_method": "protected-atp candidate preview",
        },
    )


def prune_pareto_candidates(problem: ProblemData) -> ProblemData:
    """Remove candidates dominated on estimated fill, value, cost, and lead time."""

    candidates = problem.candidates.copy()
    order_group = problem.orders.set_index("order_id").get("assignment_group")
    if bool(problem.metadata.get("enforce_assignment_group", False)) and order_group is not None:
        candidates["assignment_group"] = candidates["order_id"].map(order_group)
        option = (
            candidates.groupby(["assignment_group", "group_option_id"], as_index=False)
            .agg(
                estimated_fill_cases=("estimated_fill_cases", "sum"),
                estimated_fulfilled_value=("estimated_fulfilled_value", "sum"),
                shipping_cost=("shipping_cost", "sum"),
                lead_time_days=("lead_time_days", "max"),
                is_default=("is_default", "max"),
            )
        )
        group_columns = ["assignment_group"]
    else:
        option = candidates.copy()
        option["assignment_group"] = option["order_id"]
        group_columns = ["assignment_group"]

    keep: set[tuple[str, str]] = set()
    for _, group in option.groupby(group_columns, sort=False):
        for index, row in group.iterrows():
            key = (str(row["assignment_group"]), str(row["group_option_id"]))
            if bool(row["is_default"]):
                keep.add(key)
                continue
            competitors = group.drop(index=index)
            dominates = (
                (competitors["estimated_fill_cases"] >= row["estimated_fill_cases"])
                & (
                    competitors["estimated_fulfilled_value"]
                    >= row["estimated_fulfilled_value"]
                )
                & (competitors["shipping_cost"] <= row["shipping_cost"])
                & (competitors["lead_time_days"] <= row["lead_time_days"])
                & (
                    (competitors["estimated_fill_cases"] > row["estimated_fill_cases"])
                    | (
                        competitors["estimated_fulfilled_value"]
                        > row["estimated_fulfilled_value"]
                    )
                    | (competitors["shipping_cost"] < row["shipping_cost"])
                    | (competitors["lead_time_days"] < row["lead_time_days"])
                )
            )
            if not dominates.any():
                keep.add(key)

    if "assignment_group" not in candidates.columns:
        candidates["assignment_group"] = candidates["order_id"].map(order_group)
        candidates["assignment_group"] = candidates["assignment_group"].fillna(
            candidates["order_id"]
        )
    retained = candidates.loc[
        candidates.apply(
            lambda row: (str(row["assignment_group"]), str(row["group_option_id"]))
            in keep,
            axis=1,
        )
    ].drop(columns=["assignment_group"])
    metadata = {
        **problem.metadata,
        "pareto_pruned": True,
        "candidate_rows_before_pruning": len(problem.candidates),
        "candidate_rows_after_pruning": len(retained),
    }
    return replace(problem, candidates=retained.reset_index(drop=True), metadata=metadata)


def limit_candidates(problem: ProblemData, count: int) -> ProblemData:
    """Retain at most ``count`` candidates per decision unit, always keeping default."""

    if count < 1:
        raise ValueError("count must be at least one")
    candidates = problem.candidates.copy()
    candidates["rank_score"] = (
        candidates["estimated_fulfilled_value"].astype(float)
        - candidates["shipping_cost"].astype(float)
    )
    order_group = problem.orders.set_index("order_id").get("assignment_group")
    if bool(problem.metadata.get("enforce_assignment_group", False)) and order_group is not None:
        candidates["assignment_group"] = candidates["order_id"].map(order_group)
        option_scores = (
            candidates.groupby(["assignment_group", "group_option_id"], as_index=False)
            .agg(rank_score=("rank_score", "sum"), is_default=("is_default", "max"))
        )
        keep: set[tuple[str, str]] = set()
        for group_id, group in option_scores.groupby("assignment_group", sort=False):
            ordered = group.sort_values(
                ["is_default", "rank_score", "group_option_id"],
                ascending=[False, False, True],
                kind="mergesort",
            ).head(count)
            keep.update((str(group_id), str(value)) for value in ordered["group_option_id"])
        retained = candidates.loc[
            candidates.apply(
                lambda row: (str(row["assignment_group"]), str(row["group_option_id"]))
                in keep,
                axis=1,
            )
        ]
    else:
        retained = (
            candidates.sort_values(
                ["order_id", "is_default", "rank_score", "candidate_id"],
                ascending=[True, False, False, True],
                kind="mergesort",
            )
            .groupby("order_id", sort=False)
            .head(count)
        )
    retained = retained.drop(columns=["rank_score", "assignment_group"], errors="ignore")
    return replace(
        problem,
        candidates=retained.reset_index(drop=True),
        metadata={**problem.metadata, "candidate_limit": count},
    )


def load_poc_problem(
    bundle_dir: str | Path,
    *,
    config: PocConfig | None = None,
    strict_bundle_audit: bool = True,
) -> ProblemData:
    """Load the supplied challenge files into a validated canonical instance."""

    settings = config or PocConfig()
    settings.validate()
    if strict_bundle_audit:
        audit_poc_bundle(bundle_dir)
    source = _build_source_tables(bundle_dir)
    lines = _canonical_lines(source["orders"])
    orders = _canonical_orders(source["orders"])
    orders["min_divert_improvement_fraction"] = (
        settings.min_divert_improvement_fraction
    )
    orders, lines = _filter_focus(
        source["orders"],
        orders,
        lines,
        focus_only=settings.focus_only,
        enforce_assignment_group=settings.enforce_assignment_group,
    )
    inventory_source = _prepare_inventory(source["inventory"])
    candidates, inventory, capacities, calendar = _candidate_and_resource_tables(
        orders,
        lines,
        inventory_source,
        source["shipping"],
        source["dock"],
        source["throughput"],
        settings,
    )
    metadata = {
        "dataset_id": "nestle-wiser-poc-readable-v1",
        "schema_version": SCHEMA_VERSION,
        "assumption_version": ASSUMPTION_VERSION,
        "currency": "source_currency",
        "quantity_unit": "cases",
        "inventory_policy": "projected_atp",
        "inventory_protection_days": settings.protection_days,
        "penalty_mode": "thresholded_cut",
        "enforce_min_divert_improvement": False,
        "min_divert_improvement_fraction": settings.min_divert_improvement_fraction,
        "min_divert_improvement_cases": settings.min_divert_improvement_cases,
        "enforce_assignment_group": settings.enforce_assignment_group,
        "pick_capacity_mode": "pallet_case"
        if settings.throughput_headroom_fraction is not None
        else "cases",
        "throughput_capacity_is_scenario": (
            settings.throughput_headroom_fraction is not None
        ),
        "throughput_headroom_fraction": settings.throughput_headroom_fraction,
        "bundle_sha256": _bundle_hash(_paths(bundle_dir).values()),
        "raw_data_export_permitted": False,
    }
    problem = normalize_problem_data(
        ProblemData(
            orders=orders,
            order_lines=lines,
            inventory=inventory,
            candidates=candidates,
            capacities=capacities,
            calendar=calendar,
            metadata=metadata,
            source_dir=Path(bundle_dir),
        )
    )
    problem = _with_default_reference(problem)
    if settings.pareto_prune:
        problem = prune_pareto_candidates(problem)
    return normalize_problem_data(problem)


def subset_problem(
    problem: ProblemData,
    order_ids: Iterable[str],
    *,
    recompute_default_reference: bool = True,
) -> ProblemData:
    """Return a self-contained order subset, expanding whole assignment groups."""

    selected = {str(value) for value in order_ids}
    if not selected:
        raise ValueError("order_ids must be nonempty")
    unknown = selected - set(problem.orders["order_id"].astype(str))
    if unknown:
        raise ValueError(f"Unknown order_ids: {sorted(unknown)[:5]}")
    if bool(problem.metadata.get("enforce_assignment_group", False)):
        groups = set(
            problem.orders.loc[
                problem.orders["order_id"].astype(str).isin(selected),
                "assignment_group",
            ].astype(str)
        )
        selected = set(
            problem.orders.loc[
                problem.orders["assignment_group"].astype(str).isin(groups),
                "order_id",
            ].astype(str)
        )
    orders = problem.orders.loc[
        problem.orders["order_id"].astype(str).isin(selected)
    ].copy()
    lines = problem.order_lines.loc[
        problem.order_lines["order_id"].astype(str).isin(selected)
    ].copy()
    candidates = problem.candidates.loc[
        problem.candidates["order_id"].astype(str).isin(selected)
    ].copy()
    relevant_skus = set(lines["sku_id"].astype(str))
    relevant_dcs = set(candidates["dc_id"].astype(str))
    inventory = problem.inventory.loc[
        problem.inventory["sku_id"].astype(str).isin(relevant_skus)
        & problem.inventory["dc_id"].astype(str).isin(relevant_dcs)
    ].copy()
    candidate_keys = set(
        zip(candidates["dc_id"].astype(str), pd.to_datetime(candidates["pgi_date"]))
    )
    capacities = problem.capacities.loc[
        problem.capacities.apply(
            lambda row: (str(row["dc_id"]), pd.Timestamp(row["date"]))
            in candidate_keys,
            axis=1,
        )
    ].copy()
    calendar = problem.calendar.loc[
        problem.calendar.apply(
            lambda row: (str(row["dc_id"]), pd.Timestamp(row["date"]))
            in candidate_keys,
            axis=1,
        )
    ].copy()
    subset = normalize_problem_data(
        replace(
            problem,
            orders=orders.reset_index(drop=True),
            order_lines=lines.reset_index(drop=True),
            inventory=inventory.reset_index(drop=True),
            candidates=candidates.reset_index(drop=True),
            capacities=capacities.reset_index(drop=True),
            calendar=calendar.reset_index(drop=True),
            metadata={
                **problem.metadata,
                "dataset_id": f"{problem.metadata.get('dataset_id')}::subset-{len(selected)}",
                "enforce_min_divert_improvement": False,
            },
        )
    )
    return _with_default_reference(subset) if recompute_default_reference else subset


def select_shortage_subset(problem: ProblemData, count: int) -> ProblemData:
    """Select high-shortage decision units deterministically for tractable studies."""

    if count <= 0:
        raise ValueError("count must be positive")
    demand = problem.order_lines.groupby("order_id")["demand_cases"].sum()
    reference = problem.orders.set_index("order_id")["default_fillable_cases"]
    shortage = (demand - reference).sort_values(ascending=False, kind="mergesort")
    return subset_problem(problem, shortage.head(count).index.astype(str))


def audit_poc_outputs(bundle_dir: str | Path, problem: ProblemData | None = None) -> dict[str, object]:
    """Return privacy-safe reconciliation metrics for supplied recommendation outputs."""

    source = _build_source_tables(bundle_dir)
    order_output = source["order_output"]
    sku_output = source["sku_output"]
    order_ids = order_output["SalesDocument/GroupingIndicator"].map(_id)
    sku_order_ids = sku_output["SalesDocument/GroupingIndicator"].map(_id)
    key_match = set(order_ids) == set(sku_order_ids)
    load_ids = order_output["LoadNumber"].map(_id)
    assignment_groups = load_ids.copy()
    missing_load = assignment_groups.eq("")
    assignment_groups.loc[missing_load] = "ORDER::" + order_ids.loc[missing_load]
    selected_cases = np.where(
        order_output["IsDivert"].astype(str).eq("Non-Default"),
        _numeric(order_output["Divert_Qty_Cases"]),
        _numeric(order_output["Default_Qty_Cases"]),
    )
    result: dict[str, object] = {
        "orders": int(order_ids.nunique()),
        "order_sku_rows": len(sku_output),
        "named_loads": int(load_ids.loc[load_ids.ne("")].nunique()),
        "assignment_groups": int(assignment_groups.nunique()),
        "diverted_orders": int((order_output["IsDivert"] == "Non-Default").sum()),
        "order_sku_key_coverage": bool(key_match),
        "requested_cases": round(_numeric(order_output["OrderedQty_Cases"]).sum()),
        "selected_fulfilled_cases": round(float(selected_cases.sum())),
    }
    requested = float(result["requested_cases"])
    result["selected_case_fill_rate"] = (
        float(result["selected_fulfilled_cases"]) / requested if requested else 1.0
    )

    if problem is not None:
        default_penalty = order_output.assign(order_id=order_ids).set_index("order_id")[
            "PenaltyIfNotDiverted"
        ]
        source_lines = _canonical_lines(source["orders"])
        default_sku = sku_output.copy()
        default_sku["order_id"] = sku_order_ids
        default_sku["sku_id"] = default_sku["MaterialNumber"].map(_id)
        default_fill = default_sku.set_index(["order_id", "sku_id"])[
            "Default_Qty_Cases"
        ].map(float)
        merged = source_lines.copy()
        merged["unfulfilled_cases"] = merged.apply(
            lambda row: int(row["demand_cases"])
            - round(default_fill.loc[(row["order_id"], row["sku_id"])]),
            axis=1,
        )
        full_problem = replace(
            problem,
            orders=_canonical_orders(source["orders"]),
            order_lines=source_lines,
        )
        errors: list[float] = []
        for order_id, group in merged.groupby("order_id", sort=False):
            calculated = order_penalty(full_problem, str(order_id), group)
            errors.append(abs(calculated - float(default_penalty.loc[str(order_id)])))
        result["default_penalty_max_abs_error"] = max(errors, default=0.0)
    return result
